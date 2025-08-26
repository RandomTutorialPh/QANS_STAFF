import os, sys
import ctypes

# 🔇 Suppress ZBar (C-level) warnings by redirecting stderr
sys.stderr = open(os.devnull, 'w')

import tkinter as tk
from tkinter import ttk
import os
from tkinter import messagebox
from datetime import datetime
import qrcode
import cv2
from pyzbar import pyzbar
from threading import Thread
from pyzbar.pyzbar import decode
import time
import os
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import requests
import winsound


# ✅ Global variables for Android SMS Gateway
ANDROID_IP = None
ANDROID_TOKEN = None

# ---------------- FUNCTIONS ----------------

def clear_content():
    for widget in content_frame.winfo_children():
        widget.destroy()

def show_home():
    clear_content()

    home_frame = ttk.Frame(content_frame)
    home_frame.place(relx=0.5, rely=0.5, anchor="center")

    # Logo
    logo_path = os.path.join(IMG_DIR, "logo.png")
    if os.path.exists(logo_path):
        logo_img = tk.PhotoImage(file=logo_path)
        home_frame.logo_img = logo_img
        logo_label = ttk.Label(home_frame, image=logo_img)
        logo_label.pack(pady=10)

    title = ttk.Label(home_frame, text="QR ATTENDANCE SYSTEM", font=("Arial", 20, "bold"))
    title.pack(pady=10)

    btn_generator = ttk.Button(home_frame, text="QR Generator", command=show_form)
    btn_generator.pack(pady=10, ipadx=20, ipady=10)

    btn_scanner = ttk.Button(home_frame, text="QR Scanner", command=open_qr_scanner)
    btn_scanner.pack(pady=10, ipadx=20, ipady=10)


def show_form():
    clear_content()

    form_frame = ttk.Frame(content_frame)
    form_frame.place(relx=0.5, rely=0.5, anchor="center")

    entries = {}

    # First Name
    ttk.Label(form_frame, text="First Name:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
    entries["First Name"] = ttk.Entry(form_frame, width=40)
    entries["First Name"].grid(row=0, column=1, padx=5, pady=5)

    # Middle Name
    ttk.Label(form_frame, text="Middle Name:").grid(row=1, column=0, sticky="w", padx=5, pady=5)
    entries["Middle Name"] = ttk.Entry(form_frame, width=40)
    entries["Middle Name"].grid(row=1, column=1, padx=5, pady=5)

    # Last Name
    ttk.Label(form_frame, text="Last Name:").grid(row=2, column=0, sticky="w", padx=5, pady=5)
    entries["Last Name"] = ttk.Entry(form_frame, width=40)
    entries["Last Name"].grid(row=2, column=1, padx=5, pady=5)

    # Suffix (Dropdown)
    ttk.Label(form_frame, text="Suffix:").grid(row=3, column=0, sticky="w", padx=5, pady=5)
    suffix_values = ["", "Jr.", "Sr.", "II", "III", "IV"]
    suffix_cb = ttk.Combobox(form_frame, values=suffix_values, width=37, state="readonly")
    suffix_cb.grid(row=3, column=1, padx=5, pady=5)
    entries["Suffix"] = suffix_cb

    # Gender
    ttk.Label(form_frame, text="Gender:").grid(row=4, column=0, sticky="w", padx=5, pady=5)
    gender_cb = ttk.Combobox(form_frame, values=["Male", "Female", "Other"], width=37, state="readonly")
    gender_cb.grid(row=4, column=1, padx=5, pady=5)
    entries["Gender"] = gender_cb

    # Birthday (Day, Month, Year)
    ttk.Label(form_frame, text="Birthday:").grid(row=5, column=0, sticky="w", padx=5, pady=5)
    birthday_frame = ttk.Frame(form_frame)
    birthday_frame.grid(row=5, column=1, padx=5, pady=5, sticky="w")

    day_cb = ttk.Combobox(birthday_frame, values=list(range(1, 32)), width=5, state="readonly")
    day_cb.pack(side="left", padx=2)
    month_cb = ttk.Combobox(birthday_frame, values=list(range(1, 13)), width=5, state="readonly")
    month_cb.pack(side="left", padx=2)
    year_cb = ttk.Combobox(birthday_frame, values=list(range(1900, datetime.now().year + 1)), width=7, state="readonly")
    year_cb.pack(side="left", padx=2)

    entries["Birthday"] = (day_cb, month_cb, year_cb)

    # Address
    ttk.Label(form_frame, text="Address:").grid(row=6, column=0, sticky="w", padx=5, pady=5)
    entries["Address"] = ttk.Entry(form_frame, width=40)
    entries["Address"].grid(row=6, column=1, padx=5, pady=5)

    # Department
    ttk.Label(form_frame, text="Department:").grid(row=7, column=0, sticky="w", padx=5, pady=5)
    entries["Department"] = ttk.Entry(form_frame, width=40)
    entries["Department"].grid(row=7, column=1, padx=5, pady=5)

    # Position
    ttk.Label(form_frame, text="Position:").grid(row=8, column=0, sticky="w", padx=5, pady=5)
    entries["Position"] = ttk.Entry(form_frame, width=40)
    entries["Position"].grid(row=8, column=1, padx=5, pady=5)

    # Emergency Name
    ttk.Label(form_frame, text="Emergency Contact Name:").grid(row=9, column=0, sticky="w", padx=5, pady=5)
    entries["Emergency Contact Name"] = ttk.Entry(form_frame, width=40)
    entries["Emergency Contact Name"].grid(row=9, column=1, padx=5, pady=5)

    # Emergency Contact Number (only numbers)
    ttk.Label(form_frame, text="Emergency Contact Number:").grid(row=10, column=0, sticky="w", padx=5, pady=5)

    def validate_number(P):
        return P.isdigit() or P == ""

    vcmd = (form_frame.register(validate_number), "%P")
    entries["Emergency Contact Number"] = ttk.Entry(form_frame, width=40, validate="key", validatecommand=vcmd)
    entries["Emergency Contact Number"].grid(row=10, column=1, padx=5, pady=5)

    
    # ✅ Get base directory depending on exe or script
    if getattr(sys, 'frozen', False):  
        BASE_DIR = os.path.dirname(sys.executable)
    else:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))

    # Submit Button
    def submit_form():
        # Collect values
        first_name = entries["First Name"].get().strip().upper()
        middle_name = entries["Middle Name"].get().strip().upper()
        last_name = entries["Last Name"].get().strip().upper()
        suffix = entries["Suffix"].get().strip().upper()
        gender = entries["Gender"].get().strip().upper()

        d, m, y = entries["Birthday"]
        birthday = f"{d.get()}/{m.get()}/{y.get()}".strip().upper()

        address = entries["Address"].get().strip().upper()
        department = entries["Department"].get().strip().upper()
        position = entries["Position"].get().strip().upper()
        emergency_name = entries["Emergency Contact Name"].get().strip().upper()
        emergency_number = entries["Emergency Contact Number"].get().strip().upper()

        # Check required fields
        if not first_name or not last_name or not gender or not d.get() or not m.get() or not y.get() \
        or not address or not department or not position or not emergency_name or not emergency_number:
            messagebox.showwarning("Missing Fields", "⚠️ Please fill in all required fields before submitting.")
            return

        # If all filled, print values
        print("Form Data Submitted:")
        print("First Name:", first_name)
        print("Middle Name:", middle_name)
        print("Last Name:", last_name)
        print("Suffix:", suffix)
        print("Gender:", gender)
        print("Birthday:", birthday)
        print("Address:", address)
        print("Department:", department)
        print("Position:", position)
        print("Emergency Contact Name:", emergency_name)
        print("Emergency Contact Number:", emergency_number)
        

        # Embed hidden key in QR data
        qr_data = f"""
        FIRST NAME: {first_name}
        MIDDLE NAME: {middle_name}
        LAST NAME: {last_name}
        SUFFIX: {suffix}
        GENDER: {gender}
        BIRTHDAY: {birthday}
        ADDRESS: {address}
        DEPARTMENT: {department}
        POSITION: {position}
        EMERGENCY CONTACT NAME: {emergency_name}
        EMERGENCY CONTACT NO: {emergency_number}
        """

        # Generate QR code
        qr = qrcode.make(qr_data)

        # ✅ Save in "qrcodes" folder beside the exe/script
        save_dir = os.path.join(BASE_DIR, "qrcodes")
        os.makedirs(save_dir, exist_ok=True)

        filename = f"{first_name}_{last_name}_qr.png"
        file_path = os.path.join(save_dir, filename)

        qr.save(file_path)
        print(f"✅ QR code generated and saved at: {file_path}")

        # Success message
        messagebox.showinfo("Success", f"QR successfully generated and saved at:\n{file_path}")



        show_form()

    btn_submit = ttk.Button(form_frame, text="Generate QR", command=submit_form)
    btn_submit.grid(row=11, column=0, columnspan=2, pady=15, ipadx=10, ipady=5)

    # Home Button
    btn_home = ttk.Button(form_frame, text="🏠 Home", command=show_home)
    btn_home.grid(row=12, column=0, columnspan=2, pady=5, ipadx=10, ipady=5)
#---------------------------------------------------------------------------------------------------

# ----------------------------------#
# QR Code Scanner Functionality
# ----------------------------------#

# ✅ Try to import pygrabber for real camera names
try:
    from pygrabber.dshow_graph import FilterGraph
except ImportError:
    FilterGraph = None

# ✅ Pyzbar fallback check
try:
    from pyzbar.pyzbar import decode as pyzbar_decode
    _HAS_PYZBAR = True
except Exception:
    _HAS_PYZBAR = False


def list_cameras():
    """Return available cameras as [(index, name), ...]"""
    cameras = []
    if FilterGraph:  # if pygrabber is installed
        graph = FilterGraph()
        devices = graph.get_input_devices()
        for i, name in enumerate(devices):
            cameras.append((i, name))  # (index, device_name)
    else:
        # fallback: just check indexes
        for i in range(5):
            cap = cv2.VideoCapture(i)
            if cap.isOpened():
                cameras.append((i, f"Camera {i}"))
                cap.release()
    return cameras


def decode_with_pyzbar(frame):
    results = []
    try:
        for obj in pyzbar_decode(frame):
            data = obj.data.decode('utf-8', errors='ignore')
            points = obj.polygon
            bbox = [(p.x, p.y) for p in points] if points else None
            results.append({'data': data, 'bbox': bbox})
    except Exception:
        pass
    return results


def decode_with_opencv(frame):
    detector = cv2.QRCodeDetector()
    data, points, _ = detector.detectAndDecode(frame)
    results = []
    if points is not None and data:
        bbox = [(int(x), int(y)) for x, y in points.reshape(-1, 2)]
        results.append({'data': data, 'bbox': bbox})
    return results


def draw_bbox(frame, bbox, label):
    if not bbox:
        return
    for i in range(len(bbox)):
        pt1 = bbox[i]
        pt2 = bbox[(i + 1) % len(bbox)]
        cv2.line(frame, pt1, pt2, (0, 255, 0), 2)
    x, y = bbox[0]
    cv2.putText(frame, label, (x, max(0, y - 10)),
                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

# --- Scan flow control ---
COOLDOWN_SEC = 4
_last_scans = {}

def parse_qr_text(data: str):
    person = {}
    for line in data.splitlines():
        if ":" in line:
            k, v = line.split(":", 1)
            person[k.strip().upper()] = v.strip()
    return person

# Generate the email message    
def generate_message(action, first_name, last_name, suffix, deparment, position, current_time):
    """Generate attendance message with greeting based on AM/PM"""
    greeting = "Good morning" if datetime.now().strftime("%p") == "AM" else "Good afternoon"
    subject = f"{action.title()} Notification - {first_name} {last_name} {suffix}"
    body = (
        f"{greeting}!\n\n"
        f"This is to inform you that {first_name} {last_name} {suffix}\n"
        f"has successfully {action} at {current_time} today at Philippine Army duty.\n\n"
        "Regards,\nQR ATTENDANCE NOTIFICATION SYSTEM"
    )
    return subject, body

# ---------------- EMAIL ----------------
def send_email_notification(subject, body, recipient_email):
    sender_email = "jtechnotificationsystem@gmail.com" 
    app_password = "ryopnmnkwfreltaX"

    try:
        # Setup message
        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = recipient_email
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        # Connect to Gmail
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, app_password)
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()

        print("✅ Email sent successfully to", recipient_email)
    except Exception as e:
        print("🚨 Error sending Email:", str(e))

def send_sms(contact_no,message):
    global ANDROID_IP, ANDROID_TOKEN
    if not ANDROID_IP or not ANDROID_TOKEN:
        print("🚨 Android IP or Token not set!")
        return False

    # 📌 Static number (change this to your target number)
    contact_no_static = "09187705134"

    data = {
        "token": ANDROID_TOKEN,
        "number": contact_no_static,
        "message": message
    }

    try:
        response = requests.post(f"http://{ANDROID_IP}:5000/send_sms", data=data, timeout=10)
        print("📨 SMS Gateway Status:", response.status_code, response.text)
        if response.status_code == 200:
            print("✅ SMS sent successfully via Termux")
            return True
        else:
            print("❌ Failed to send SMS:", response.text)
            return False
    except Exception as e:
        print("🚨 Error sending SMS:", str(e))
        return False

def notify_user(subject, body, emergency_contact, guardian_email=None):
    """Try SMS first. If successful, send email. If SMS fails, still send email."""
    sms_ok = False

    if emergency_contact:
        print("📱 Trying to send SMS...")
        try:
            send_sms(emergency_contact, body)  # your Termux SMS function
            sms_ok = True
        except Exception as e:
            print("❌ SMS failed:", e)

    if guardian_email:
        if sms_ok:
            print("✅ SMS success, now sending Email...")
        else:
            print("⚠️ SMS failed, sending Email instead...")

        send_email_notification(subject, body, guardian_email)

def process_backend_and_format_message(person_data):
    first_name = person_data.get("FIRST NAME", "")
    middle_name = person_data.get("MIDDLE NAME", "")
    last_name = person_data.get("LAST NAME", "")
    suffix = person_data.get("SUFFIX", "")
    gender = person_data.get("GENDER", "")
    birthday = person_data.get("BIRTHDAY", "")
    address = person_data.get("ADDRESS", "")
    department = person_data.get("DEPARTMENT", "")
    position = person_data.get("POSITION", "")
    emergency_name = person_data.get("EMERGENCY CONTACT NAME", "")
    emergency_contact = person_data.get("EMERGENCY CONTACT NO", "")

    now = datetime.now()
    time_out = now.strftime("%I:%M %p")
    date_str = now.strftime("%Y-%m-%d")
    time_in = now.strftime("%I:%M%p").lower()

    if getattr(sys, 'frozen', False):
        base_path = os.path.dirname(sys.executable)
    else:
        base_path = os.path.dirname(__file__)

    folder_path = os.path.join(base_path, "attendance", "SUMMARY OF ATTENDANCE")
    os.makedirs(folder_path, exist_ok=True)
    file_name = f"{date_str}_attendance_record.xlsx"
    file_path = os.path.join(folder_path, file_name)

    headers = [
        "FIRST NAME", "MIDDLE NAME", "LAST NAME", "SUFFIX", "GENDER",
        "BIRTHDAY", "ADDRESS", "DEPARTMENT", "POSITION", "TIME IN", "TIME OUT",
        "EMERGENCY CONTACT NAME", "EMERGENCY CONTACT NO", "DATE"
    ]

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(headers)
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    person_found = False
    did_timeout = False

    for row in ws.iter_rows(min_row=2, values_only=False):
        row_fname = (row[0].value or '').strip().upper()
        row_mname = (row[1].value or '').strip().upper()
        row_lname = (row[2].value or '').strip().upper()
        row_suffix = (row[3].value or '').strip().upper()

        if (row_fname == first_name and row_mname == middle_name and
            row_lname == last_name and row_suffix == suffix):
            if not row[10].value:
                row[10].value = time_out
                did_timeout = True
                subject, body = generate_message("timed out", first_name, last_name, suffix, department, position, time_out)
                # notify_user(subject, body, emergency_contact, guardian_email="randomtutorialph@gmail.com")
                notify_user(subject, body, emergency_contact, guardian_email="mayojanel066@gmail.com")
            person_found = True
            break

    if not person_found:
        new_row = [
            first_name, middle_name, last_name, suffix, gender, birthday,
            address, department, position, time_in, '', emergency_name, emergency_contact, date_str
        ]
        ws.append(new_row)
        subject, body = generate_message("timed in", first_name, last_name, suffix, department, position, time_out)
        #notify_user(subject, body, emergency_contact, guardian_email="randomtutorialph@gmail.com")
        notify_user(subject, body, emergency_contact, guardian_email="mayojanel066@gmail.com")

    wb.save(file_path)
    action_text = "timed out" if did_timeout else "timed in"
    ui_text = f"{first_name} {last_name} successfully {action_text}."
    return ui_text, action_text

def show_processing_screen(msg="Please wait…"):
    clear_content()
    frame = ttk.Frame(content_frame)
    frame.place(relx=0.5, rely=0.5, anchor="center")
    ttk.Label(frame, text=msg, font=("Arial", 16, "bold")).pack(pady=10)
    return frame

def show_temporary_status_then_reopen(status_text, camera_name, delay_ms=2000):
    show_processing_screen(status_text)
    root.after(delay_ms, lambda: open_qr_scanner_with_camera(camera_name))

def open_qr_scanner_with_camera(camera_name):
    clear_content()
    scanner_frame = ttk.Frame(content_frame)
    scanner_frame.place(relx=0.5, rely=0.5, anchor="center")
    ttk.Label(scanner_frame, text="📷 QR Scanner", font=("Arial", 20)).pack(pady=10)
    ttk.Label(scanner_frame, text=f"Using: {camera_name}").pack(pady=5)
    start_scanner(camera_name)
    ttk.Button(scanner_frame, text="🏠 Home", command=show_home).pack(pady=10)

def handle_scan_then_backend(qr_text, selected_camera):
    # Show success + data UI first (and destroy current UI)
    show_scan_success_ui(qr_text)

    # Now do backend in background, then bounce back to scanner
    def _worker():
        try:
            person = parse_qr_text(qr_text)
            status_text, _ = process_backend_and_format_message(person)
        except Exception as e:
            status_text = f"Error: {e}"
        # After processing, show a brief status then reopen the scanner with same camera
        root.after(0, lambda: show_temporary_status_then_reopen(status_text, selected_camera, delay_ms=2000))

    Thread(target=_worker, daemon=True).start()


def start_scanner(selected_camera):
    import warnings
    warnings.filterwarnings("ignore")  # hide pyzbar warnings

    cap = None
    for i, name in list_cameras():
        if name == selected_camera:
            cap = cv2.VideoCapture(i)
            break

    if cap is None or not cap.isOpened():
        messagebox.showerror("Error", "Cannot open selected camera")
        return

    last_print = time.time()

    while True:
        ok, frame = cap.read()
        if not ok:
            break

        # Try pyzbar first
        results = []
        try:
            if _HAS_PYZBAR:
                results = decode_with_pyzbar(frame)
        except Exception:
            results = []

        # Fallback to OpenCV if pyzbar found nothing
        if not results:
            results = decode_with_opencv(frame)

        # Draw and handle results
        for r in results:
            data = r['data']
            bbox = r['bbox']
            draw_bbox(frame, bbox, data[:50] + ("..." if len(data) > 50 else ""))

            nowt = time.time()
            last_t = _last_scans.get(data, 0)
            if nowt - last_t < COOLDOWN_SEC:
                continue  # still cooling down for this QR

            _last_scans[data] = nowt

            if data:
                # Debug + audible cue
                print(f"[DEBUG] QR Detected: {data[:100]}{'...' if len(data) > 100 else ''}")
                winsound.Beep(1000, 200)

                # ✅ CLOSE/RELEASE CAMERA IMMEDIATELY and break out of the loop
                cap.release()
                cv2.destroyAllWindows()

                # ✅ Destroy current UI and show the scanned result UI,
                #    then process backend, then reopen scanner with same camera
                handle_scan_then_backend(data, selected_camera)

                # Break out of the capture loop so we don't keep reading frames
                return

        cv2.imshow("QR Scanner - press Q to quit", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            print("👋 Scanner stopped by user.")
            break

        # Occasional console heartbeat
        if time.time() - last_print > 15:
            print("Scanning... (press 'q' to quit)")
            last_print = time.time()

    # Clean up
    cap.release()
    cv2.destroyAllWindows()

    # After quitting, go back to camera selection screen
    root.after(100, open_qr_scanner)

def show_scan_success_ui(qr_text: str):
    """Show a success screen with the raw QR data (nicely parsed) before backend work."""
    clear_content()
    frame = ttk.Frame(content_frame)
    frame.place(relx=0.5, rely=0.5, anchor="center")

    ttk.Label(frame, text="✅ QR Scanned Successfully!", font=("Arial", 18, "bold")).pack(pady=(0, 12))

    # Parse and display fields
    parsed = parse_qr_text(qr_text)
    box = ttk.Frame(frame)
    box.pack(pady=6)

    # Show key/value lines (hide completely empty values)
    for k in [
        "FIRST NAME","MIDDLE NAME","LAST NAME","SUFFIX","GENDER","BIRTHDAY",
        "ADDRESS","DEPARTMENT","POSITION","EMERGENCY CONTACT NAME","EMERGENCY CONTACT NO"
    ]:
        val = parsed.get(k, "")
        if val:
            row = ttk.Frame(box)
            row.pack(anchor="w")
            ttk.Label(row, text=f"{k}:", font=("Arial", 10, "bold")).pack(side="left")
            ttk.Label(row, text=f" {val}", font=("Arial", 10)).pack(side="left")

    # A small note while the backend runs
    ttk.Label(frame, text="Processing attendance…", font=("Arial", 11)).pack(pady=(10, 0))



def open_qr_scanner():
    clear_content()

    # --- Step 1: Ask for Android SMS Gateway info
    input_frame = ttk.Frame(content_frame)
    input_frame.place(relx=0.5, rely=0.5, anchor="center")

    ttk.Label(input_frame, text="📱 Enter Android SMS Gateway Details", font=("Arial", 14, "bold")).pack(pady=10)

    # IP Address
    ttk.Label(input_frame, text="Android IP Address:").pack(pady=5)
    ip_entry = ttk.Entry(input_frame, width=30)
    ip_entry.pack(pady=5)

    # Token
    ttk.Label(input_frame, text="Token:").pack(pady=5)
    token_entry = ttk.Entry(input_frame, width=30, show="*")
    token_entry.pack(pady=5)

    def save_and_open():
        global ANDROID_IP, ANDROID_TOKEN
        ANDROID_IP = ip_entry.get().strip()
        ANDROID_TOKEN = token_entry.get().strip()

        if not ANDROID_IP or not ANDROID_TOKEN:
            messagebox.showwarning("Missing", "⚠️ Please enter both IP and Token.")
            return

        # --- Step 2: Go to camera selection
        clear_content()
        
        scanner_frame = ttk.Frame(content_frame)
        scanner_frame.place(relx=0.5, rely=0.5, anchor="center")
        ttk.Label(scanner_frame, text="📷 QR Scanner", font=("Arial", 20)).pack(pady=10)
        cameras = list_cameras()
        if not cameras:
            ttk.Label(scanner_frame, text="No cameras found").pack(pady=10)
            return
        camera_var = tk.StringVar(value=cameras[0][1])
        camera_dropdown = ttk.Combobox(
            scanner_frame, textvariable=camera_var,
            values=[name for (_, name) in cameras], state="readonly", width=40
        )
        camera_dropdown.pack(pady=5)
        ttk.Button(scanner_frame, text="Start Scanning",
                   command=lambda: start_scanner(camera_var.get())).pack(pady=10)
        ttk.Button(scanner_frame, text="🏠 Home", command=show_home).pack(pady=10)

    ttk.Button(input_frame, text="Proceed to Scanner", command=save_and_open).pack(pady=15)
    ttk.Button(input_frame, text="🏠 Home", command=show_home).pack(pady=5)

# ---------------- MAIN ----------------
root = tk.Tk()
root.title("QR Attendance Notification System")
root.geometry("600x700")

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IMG_DIR = os.path.join(BASE_DIR, "src", "img")

icon_path = os.path.join(IMG_DIR, "icon.ico")
if os.path.exists(icon_path):
    root.iconbitmap(icon_path)

content_frame = ttk.Frame(root)
content_frame.pack(fill="both", expand=True)

show_home()

root.mainloop()
