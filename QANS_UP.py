import tkinter as tk
import serial
import threading
import time
import os
import sys
import qrcode
from tkinter import messagebox
from datetime import datetime
from PIL import Image, ImageTk
from serial.tools import list_ports
from openpyxl import Workbook, load_workbook



# ================= GLOBAL VARIABLE ===========
SELECTED_SMS_PORT = None

# ================= FUNCTIONS =================
def show_home():
    home_frame.tkraise()

def start_scan_mode():
    scan_frame.tkraise()
    scanned_data_label.config(text="Waiting for scan...")
    scan_entry.delete(0, tk.END)
    scan_entry.focus()

def show_generate_qr():
    generate_qr_frame.tkraise()

def scan_qr():
    select_com_and_scan()

def select_com_and_scan():
    popup = tk.Toplevel(root)
    popup.title("Select COM Port")
    popup.geometry("400x300")
    popup.resizable(False, False)
    popup.grab_set()  # modal

    tk.Label(
        popup,
        text="Select COM Port for SMS",
        font=("Arial", 14, "bold")
    ).pack(pady=20)

    com_var = tk.StringVar(value="COM4")

    com_frame = tk.Frame(popup)
    com_frame.pack(pady=10)

    # Scan available COM ports dynamically
    ports = list_ports.comports()

    valid_ports = [
        p for p in ports
        if "AT" in p.description.upper() or "SIMCOM" in p.description.upper()
    ]

    if not valid_ports:
        tk.Label(
            com_frame,
            text="No SIM7600 modem detected",
            font=("Arial", 12),
            fg="red"
        ).pack()
    else:
        for port in valid_ports:
            tk.Radiobutton(
                com_frame,
                text=f"{port.device}  ({port.description})",
                variable=com_var,
                value=port.device,
                font=("Arial", 11)
            ).pack(anchor="w")

        com_var.set(valid_ports[0].device)



    def confirm_and_scan():
        global SELECTED_SMS_PORT
        SELECTED_SMS_PORT = com_var.get()
        popup.destroy()
        start_scan_mode()

    tk.Button(
        popup,
        text="Scan Now",
        command=confirm_and_scan,
        bg="#2563EB",
        fg="white",
        font=("Arial", 12, "bold"),
        width=15,
        height=2
    ).pack(pady=20)


def reset_scan_mode():
    scanned_data_label.config(text="Waiting for scan...")
    scan_entry.focus()

def record_attendance(fname, mname, lname, suffix):
    attendance_dir = os.path.join(BASE_DIR, "Attendance")
    os.makedirs(attendance_dir, exist_ok=True)

    today_str = datetime.now().strftime("%m-%d-%y")
    file_path = os.path.join(attendance_dir, f"{today_str}_attendance.xlsx")

    headers = [
        "FIRST NAME", "MIDDLE NAME", "LAST NAME", "SUFFIX",
        "TIME IN AM", "TIME OUT AM",
        "TIME IN PM", "TIME OUT PM"
    ]

    # Create file if not exists
    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(headers)
        wb.save(file_path)

    wb = load_workbook(file_path)
    ws = wb.active

    now = datetime.now()
    time_str = now.strftime("%I:%M %p")
    hour = now.hour

    # Find existing record
    for row in ws.iter_rows(min_row=2):
        if (
            row[0].value == fname and
            row[1].value == mname and
            row[2].value == lname and
            row[3].value == suffix
        ):
            # AM logic
            if hour < 12 and not row[4].value:
                row[4].value = time_str
            # PM logic (future-ready)
            elif hour >= 12 and not row[6].value:
                row[6].value = time_str

            wb.save(file_path)
            return

    # New record (first scan)
    new_row = [fname, mname, lname, suffix, None, None, None, None]

    if hour < 12:
        new_row[4] = time_str  # TIME IN AM
    else:
        new_row[6] = time_str  # TIME IN PM

    ws.append(new_row)
    wb.save(file_path)


def process_scan():
    scanned_data = scan_entry.get().strip()
    if not scanned_data:
        return

    try:
        fname, mname, lname, suffix, address, guardian_name, guardian_contact = scanned_data.split("|")

        full_name = f"{fname} {mname} {lname}".replace(" . ", " ").strip()
        if suffix != ".":
            full_name += f" {suffix}"

        # ✅ RECORD EXCEL HERE (BEFORE SMS)
        record_attendance(fname, mname, lname, suffix)


    except ValueError:
        scanned_data_label.config(text="Invalid QR data!")
        scan_entry.delete(0, tk.END)
        scan_entry.focus()
        return

    # Convert guardian number to international format
    if guardian_contact.startswith("09"):
        phone = "+63" + guardian_contact[1:]
    else:
        phone = guardian_contact

    sms_message = (
        f"ALERT:\n"
        f"{full_name} has arrived safely.\n"
        f"Location: {address}"
    )

    # Update UI: show info + SMS pending
    scanned_data_label.config(
        text=(
            "📋 Personal Data\n"
            f"Name: {full_name}\n"
            f"Address: {address}\n\n"
            "🚨 In case of Emergency\n"
            f"Name: {guardian_name}\n"
            f"Contact: {guardian_contact}\n\n"
            "📨 Sending SMS..."
        )
    )

    scan_entry.delete(0, tk.END)

    # Callback after SMS attempt
    def on_sms_done(success, response):
        if success:
            scanned_data_label.config(
                text=scanned_data_label.cget("text") + "\n\n✅ SMS sent successfully!"
            )
        else:
            scanned_data_label.config(
                text=scanned_data_label.cget("text") + "\n\n❌ SMS failed!"
            )

        # Return to scan mode after 5 seconds
        scan_frame.after(5000, reset_scan_mode)

    send_sms(phone, sms_message, callback=on_sms_done)



def open_folder():
    folder_path = BASE_DIR

    if sys.platform.startswith("win"):
        os.startfile(folder_path)
    elif sys.platform == "darwin":
        os.system(f"open '{folder_path}'")
    else:
        os.system(f"xdg-open '{folder_path}'")

def only_numbers(new_value):
    return new_value.isdigit() or new_value == ""

def contact_number_validation(new_value):
    return (new_value.isdigit() and len(new_value) <= 11) or new_value == ""

def generate_qr():
    fname = entry_fname.get().strip().upper()
    mname = entry_mname.get().strip().upper()
    lname = entry_lname.get().strip().upper()
    suffix = suffix_var.get().strip().upper()
    address = entry_address.get().strip().upper()

    guardian_name = entry_guardian_name.get().strip().upper()
    guardian_contact = entry_guardian_contact.get().strip()

    # Validation
    if not fname or not mname or not lname or not address or not guardian_name or not guardian_contact:
        messagebox.showerror("Validation Error", "All fields are required.\nUse '.' if no middle name.")
        return

    if mname == "":
        messagebox.showerror("Validation Error", "Middle name required. Use '.' if none.")
        return

    if not guardian_contact.isdigit() or len(guardian_contact) != 11:
        messagebox.showerror("Validation Error", "Guardian contact must be exactly 11 digits.")
        return

    # Normalize suffix
    if suffix == "NONE":
        suffix = "."

    # QR DATA FORMAT (UPDATED)
    qr_data = f"{fname}|{mname}|{lname}|{suffix}|{address}|{guardian_name}|{guardian_contact}"

    # Save QR codes in current project directory
    qr_folder = os.path.join(BASE_DIR, "QR Code")
    os.makedirs(qr_folder, exist_ok=True)

    # Build filename: FIRST_MIDDLE_LAST_SUFFIX
    name_parts = [fname, mname, lname]

    if suffix != ".":
        name_parts.append(suffix)

    safe_name = "_".join(name_parts)
    qr_path = os.path.join(qr_folder, f"{safe_name}.png")

    qr_path = os.path.join(qr_folder, f"{safe_name}.png")

    try:
        qr = qrcode.make(qr_data)
        qr.save(qr_path)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to generate QR Code.\n{e}")
        return

    messagebox.showinfo("Success", f"QR Code generated successfully!\n\nSaved to:\n{qr_path}")

    # Clear fields
    entry_fname.delete(0, tk.END)
    entry_mname.delete(0, tk.END)
    entry_lname.delete(0, tk.END)
    entry_address.delete(0, tk.END)
    entry_guardian_name.delete(0, tk.END)
    entry_guardian_contact.delete(0, tk.END)
    suffix_var.set("NONE")


def send_sms(phone, message, callback=None):
    def worker():
        try:
            ser = serial.Serial(SELECTED_SMS_PORT, 115200, timeout=2)

            def at(cmd, wait=0.5):
                ser.write((cmd + "\r").encode())
                time.sleep(wait)
                return ser.read_all().decode(errors="ignore")

            # Reset & setup SMS
            at("AT")
            at("AT+CSCS=\"GSM\"")
            at("AT+CMGF=1")

            # Send SMS
            ser.write(f'AT+CMGS="{phone}"\r'.encode())
            time.sleep(1)
            ser.write((message + "\x1A").encode())
            time.sleep(3)

            response = ser.read_all().decode(errors="ignore")
            ser.close()

            success = "+CMGS:" in response and "OK" in response

        except Exception as e:
            success = False
            response = str(e)

        # Notify UI thread
        if callback:
            root.after(0, lambda: callback(success, response))

    threading.Thread(target=worker, daemon=True).start()


def exit_app():
    root.destroy()

# ================= BASE PATH =================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(BASE_DIR, "src", "img", "logo.png")
ICON_PATH = os.path.join(BASE_DIR, "src", "img", "logo.ico")

# ================= MAIN WINDOW =================
root = tk.Tk()

vcmd_numbers = root.register(only_numbers)
vcmd_contact = root.register(contact_number_validation)

root.overrideredirect(True)
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
root.geometry(f"{screen_width}x{screen_height}+0+0")
root.bind("<Escape>", lambda e: None)

try:
    if os.path.exists(ICON_PATH):
        root.iconbitmap(default=ICON_PATH)
except:
    pass

root.configure(bg="#f4f6f9")
root.columnconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

# ================= HEADER =================
header = tk.Frame(root, bg="#1e293b", height=100)
header.grid(row=0, column=0, sticky="nsew")
header.columnconfigure(1, weight=1)

logo = None
try:
    if os.path.exists(LOGO_PATH):
        logo_img = Image.open(LOGO_PATH).resize((70, 70))
        logo = ImageTk.PhotoImage(logo_img)
except:
    pass

if logo:
    lbl_logo = tk.Label(header, image=logo, bg="#1e293b")
    lbl_logo.image = logo
else:
    lbl_logo = tk.Label(header, bg="#1e293b", width=10)

lbl_logo.grid(row=0, column=0, padx=20, pady=10)

tk.Label(
    header,
    text="QR ATTENDANCE NOTIFICATION SYSTEM",
    font=("Arial", 26, "bold"),
    bg="#1e293b",
    fg="white"
).grid(row=0, column=1, sticky="w")

# ================= CONTENT =================
content = tk.Frame(root, bg="#f4f6f9")
content.grid(row=1, column=0, sticky="nsew")
content.columnconfigure(0, weight=1)
content.rowconfigure(0, weight=1)

# ================= HOME FRAME =================
home_frame = tk.Frame(content, bg="#f4f6f9")
home_frame.grid(row=0, column=0, sticky="nsew")

center_frame = tk.Frame(home_frame, bg="#f4f6f9")
center_frame.place(relx=0.5, rely=0.5, anchor="center")

row1 = tk.Frame(center_frame, bg="#f4f6f9")
row1.pack(pady=25)

tk.Button(
    row1, text="Generate QR Code",
    command=show_generate_qr,
    bg="#2563EB", fg="white",
    activebackground="#1D4ED8",
    font=("Arial", 16, "bold"),
    width=22, height=2, bd=0
).pack(side="left", padx=30)

tk.Button(
    row1, text="Scan QR Code",
    command=scan_qr,
    bg="#0EA5E9", fg="white",
    activebackground="#0284C7",
    font=("Arial", 16, "bold"),
    width=22, height=2, bd=0
).pack(side="left", padx=30)

row2 = tk.Frame(center_frame, bg="#f4f6f9")
row2.pack(pady=25)

tk.Button(
    row2, text="Open Folder",
    command=open_folder,
    bg="#16A34A", fg="white",
    activebackground="#15803D",
    font=("Arial", 16, "bold"),
    width=22, height=2, bd=0
).pack(side="left", padx=30)

tk.Button(
    row2, text="Exit",
    command=exit_app,
    bg="#DC2626", fg="white",
    activebackground="#991B1B",
    font=("Arial", 16, "bold"),
    width=22, height=2, bd=0
).pack(side="left", padx=30)

# ================= GENERATE QR FRAME =================
generate_qr_frame = tk.Frame(content, bg="#f4f6f9")
generate_qr_frame.grid(row=0, column=0, sticky="nsew")

# ================= SCAN QR FRAME =================
scan_frame = tk.Frame(content, bg="#f4f6f9")
scan_frame.grid(row=0, column=0, sticky="nsew")

scan_label = tk.Label(
    scan_frame,
    text="Scan QR Code",
    font=("Arial", 22, "bold"),
    bg="#f4f6f9"
)
scan_label.pack(pady=(50, 20))

# Display scanned data
scanned_data_label = tk.Label(
    scan_frame,
    text="Waiting for scan...",
    font=("Arial", 18),
    bg="#f4f6f9",
    fg="#1e293b",
    wraplength=600,
    justify="center"
)
scanned_data_label.pack(pady=20)

# Hidden entry to capture scanner input
scan_entry = tk.Entry(scan_frame)
scan_entry.pack()
scan_entry.place(x=-100, y=-100)  # keep it hidden
scan_entry.bind("<Return>", lambda e: process_scan())

# Home button for Scan screen
home_btn = tk.Button(
    scan_frame,
    text="HOME",
    command=show_home,
    bg="#6B7280",
    fg="white",
    font=("Arial", 14, "bold"),
    width=18,
    height=2,
    bd=0
)
home_btn.pack(pady=20)

# ================= GENERATE QR FORM =================
form = tk.Frame(generate_qr_frame, bg="#f4f6f9")
form.place(relx=0.5, rely=0.5, anchor="center")

tk.Label(
    form,
    text="Generate QR Code",
    font=("Arial", 22, "bold"),
    bg="#f4f6f9"
).pack(pady=(0, 30))

user_info = tk.LabelFrame(
    form,
    text=" USER INFORMATION ",
    font=("Arial", 14, "bold"),
    bg="white",
    fg="#1e293b",
    padx=30,
    pady=20
)
user_info.pack(fill="x", pady=(0, 25))

tk.Label(user_info, text="FIRST NAME:", bg="white", font=("Arial", 13))\
    .grid(row=0, column=0, sticky="w", padx=(0, 20), pady=10)
entry_fname = tk.Entry(user_info, font=("Arial", 13), width=35)
entry_fname.grid(row=0, column=1, pady=10)

tk.Label(user_info, text="MIDDLE NAME:", bg="white", font=("Arial", 13))\
    .grid(row=1, column=0, sticky="w", padx=(0, 20), pady=10)
entry_mname = tk.Entry(user_info, font=("Arial", 13), width=35)
entry_mname.grid(row=1, column=1, pady=10)

tk.Label(user_info, text="LAST NAME:", bg="white", font=("Arial", 13))\
    .grid(row=2, column=0, sticky="w", padx=(0, 20), pady=10)
entry_lname = tk.Entry(user_info, font=("Arial", 13), width=35)
entry_lname.grid(row=2, column=1, pady=10)

tk.Label(user_info, text="SUFFIX:", bg="white", font=("Arial", 13))\
    .grid(row=3, column=0, sticky="w", padx=(0, 20), pady=10)

suffix_var = tk.StringVar(value="NONE")
suffix_list = ["NONE", "JR", "SR", "II", "III", "IV"]

suffix_menu = tk.OptionMenu(user_info, suffix_var, *suffix_list)
suffix_menu.config(font=("Arial", 12), width=30)
suffix_menu.grid(row=3, column=1, pady=10)

tk.Label(user_info, text="ADDRESS:", bg="white", font=("Arial", 13))\
    .grid(row=4, column=0, sticky="w", padx=(0, 20), pady=10)
entry_address = tk.Entry(user_info, font=("Arial", 13), width=35)
entry_address.grid(row=4, column=1, pady=10)


guardian_info = tk.LabelFrame(
    form,
    text=" SMS GUARDIAN INFORMATION ",
    font=("Arial", 14, "bold"),
    bg="white",
    fg="#1e293b",
    padx=30,
    pady=20
)
guardian_info.pack(fill="x", pady=(0, 30))

tk.Label(guardian_info, text="NAME:", bg="white", font=("Arial", 13))\
    .grid(row=0, column=0, sticky="w", padx=(0, 20), pady=10)
entry_guardian_name = tk.Entry(guardian_info, font=("Arial", 13), width=35)
entry_guardian_name.grid(row=0, column=1, pady=10)

tk.Label(guardian_info, text="CONTACT NO:", bg="white", font=("Arial", 13))\
    .grid(row=1, column=0, sticky="w", padx=(0, 20), pady=10)
entry_guardian_contact = tk.Entry(
    guardian_info, font=("Arial", 13), width=35,
    validate="key", validatecommand=(vcmd_contact, "%P")
)
entry_guardian_contact.grid(row=1, column=1, pady=10)

btn_frame = tk.Frame(form, bg="#f4f6f9")
btn_frame.pack(pady=10)

tk.Button(
    btn_frame,
    text="GENERATE QR",
    command=generate_qr,
    bg="#2563EB",
    fg="white",
    font=("Arial", 14, "bold"),
    width=18,
    height=2,
    bd=0
).pack(side="left", padx=20)

tk.Button(
    btn_frame,
    text="HOME",
    command=show_home,
    bg="#6B7280",
    fg="white",
    font=("Arial", 14, "bold"),
    width=18,
    height=2,
    bd=0
).pack(side="left", padx=20)

# ================= START =================
show_home()
root.mainloop()
